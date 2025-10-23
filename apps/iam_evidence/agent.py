import os
import sys
import logging
import base64
import uuid
from pathlib import Path
from io import BytesIO
import csv
from typing import Dict, List, Optional, Any

from dotenv import load_dotenv
from google.adk.agents.llm_agent import LlmAgent
from google.adk.tools.function_tool import FunctionTool
from tachyon_adk_client import TachyonAdkClient

# Load environment variables (prefer local .env in this package)
from pathlib import Path as _Path
_env_path = _Path(__file__).parent / '.env'
if _env_path.exists():
    load_dotenv(dotenv_path=_env_path, override=True)
else:
    load_dotenv(override=True)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Add current directory to path for imports
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

# Import direct tools - try relative first, then absolute
try:
    from .database_tools import DatabaseTools
    from .jira_tools import JiraTools
    from .mongodb_tools import MongoDBTools
except ImportError:
    from database_tools import DatabaseTools
    from jira_tools import JiraTools
    from mongodb_tools import MongoDBTools

# Optional dependencies for tabular and MSG parsing
try:
    import pandas as _pd  # type: ignore
    _PANDAS_AVAILABLE = True
except Exception:
    _pd = None
    _PANDAS_AVAILABLE = False

try:
    import openpyxl as _openpyxl  # type: ignore
    _OPENPYXL_AVAILABLE = True
except Exception:
    _openpyxl = None
    _OPENPYXL_AVAILABLE = False

try:
    import extract_msg as _extract_msg  # type: ignore
    _EXTRACT_MSG_AVAILABLE = True
except Exception:
    _extract_msg = None
    _EXTRACT_MSG_AVAILABLE = False


class NHAComplianceAgent:
    """Non-MCP NHA Compliance Agent using direct tools with LLM-driven delegation"""

    def __init__(self):
        # Initialize tool instances
        mongo_tools_instance = MongoDBTools()

        # Core tools (Mongo-only, as validation is evidence-driven and DB is not required)
        self.tools = [
            FunctionTool(mongo_tools_instance.get_system_instructions),
            FunctionTool(mongo_tools_instance.analyze_evidence_files),
            FunctionTool(mongo_tools_instance.store_compliance_report),
        ]

        # Add local file ingestion tool so agents can load uploaded evidence
        self.tools.append(FunctionTool(self.load_uploaded_evidence))
        # Add native chat upload ingestion tool (accepts files from ADK Web UI)
        self.tools.append(FunctionTool(self.ingest_uploaded_evidence))
        # Add tools for CSV/Excel and Outlook MSG parsing
        self.tools.append(FunctionTool(self.convert_tabular_to_text_by_ids))
        self.tools.append(FunctionTool(self.extract_outlook_msg_by_ids))

        self.mongo_tools = mongo_tools_instance
        self._setup_environment()
        # Build a dynamic pool of control-specific sub agents (empty at boot; created on-demand)
        self.control_agents = {}
        # Map original control_id -> sanitized agent_name
        self.control_agent_name_map = {}
        # In-memory evidence cache: id -> { filename, content }
        self._evidence_cache: Dict[str, Dict[str, Any]] = {}
        # Create the coordinator agent
        self._create_agent()

    def _setup_environment(self):
        """Setup environment variables for optimal performance"""
        # Set optimal configuration
        os.environ.setdefault('SCHEMA_SOURCE', 'db')
        os.environ.setdefault('TEXT_TO_SQL_FAST_MODE', 'true')
        os.environ.setdefault('DEBUG_MODE', 'false')

        # Configure supported file types for comprehensive document processing
        default_extensions = 'pdf,doc,docx,txt,rtf,xlsx,xls,csv,jpg,jpeg,png,tiff,bmp,ppt,pptx,html,xml,json'
        os.environ.setdefault('ALLOWED_FILE_EXTENSIONS', default_extensions)
        print(f"[INFO] Supported file extensions: {os.getenv('ALLOWED_FILE_EXTENSIONS')}")

        # Local uploads directory
        os.environ.setdefault('EVIDENCE_UPLOADS_DIR', 'uploads')

    def _print_config(self):
        """Print configuration for debugging"""
        print(f"[INFO] Initializing NHA Compliance System (Non-MCP Version)")
        print(f"[CONFIG] Model: {os.getenv('MODEL', 'Not Set')}")
        print(f"[CONFIG] Agent: {os.getenv('ROOT_AGENT_NAME', 'NHA_Compliance_Assistant')}")
        print(f"[CONFIG] Database: Disabled in Non-MCP agent")
        print(f"[CONFIG] MongoDB DB: {os.getenv('MONGODB_DATABASE', 'Not Set')}")
        print(f"[CONFIG] Prompt Collection: {os.getenv('MONGODB_SYSTEM_INSTRUCTIONS_COLLECTION', 'prompt')}")

    def _create_agent(self):
        """Create the coordinator LLM agent that delegates via transfer_to_agent."""
        coordinator_name = os.getenv('ROOT_AGENT_NAME', 'NHA_Coordinator')
        self.agent = LlmAgent(
            model=TachyonAdkClient(model_name=f"openai/{os.getenv('MODEL', 'gemini-2.0-flash')}",),
            name=coordinator_name,
            description="Coordinator that routes each control to its specialized sub-agent via LLM-driven transfer.",
            instruction=(
                "You are the coordinator for NHA evidence validation.\n\n"
                "First, collect all inputs in one turn: pairs of (control_id, app_name) and a list of evidence files.\n"
                "Users may provide multiple controls.\n"
                "Immediately call the tool ensure_control_agents with the list of control_ids provided to dynamically prepare sub-agents.\n"
                "For each (control_id, app_name), transfer_to_agent(agent_name from the mapping returned by ensure_control_agents)\n"
                "Sub-agent behavior: fetch system instructions from MongoDB with prompt_name equal to the control_id; select and analyze only relevant files; then synthesize a final combined compliance report across the provided controls and apps.\n"
                "After all sub-agents finish, synthesize the final combined compliance report across the provided controls and apps.\n"
                "Rules: Do not ask for confirmations between steps. Proceed autonomously after the initial data collection. Prefer short responses."
            ),
            tools=self.tools + [FunctionTool(self.ensure_control_agents), FunctionTool(self.analyze_evidence_by_ids)],
            sub_agents=[],
        )

    def _sanitize_agent_name(self, name: str) -> str:
        import re as _re
        s = _re.sub(r"[^0-9a-zA-Z_]+", "_", name)
        if s and s[0].isdigit():
            s = f"_{s}"
        return s

    def _create_control_agent(self, control_id: str) -> LlmAgent:
        """Create or return a specialized sub-agent for a specific control ID.
        Loads system instructions from MongoDB using the control_id as the prompt name.
        """
        # Check if agent already exists
        if control_id in self.control_agents:
            return self.control_agents[control_id]

        # Load system instructions from MongoDB
        prompt_name = control_id  # per requirement, use the control ID as the variable/prompt name
        system_instruction = self.mongo_tools.get_system_instructions(prompt_name)

        if not system_instruction:
            raise ValueError(f"No system instruction found for Control ID: {control_id}")

        # Create a new agent for the specific control
        control_agent = LlmAgent(
            model=TachyonAdkClient(model_name=f"openai/{os.getenv('MODEL', 'gemini-2.0-flash')}",),
            name=self._sanitize_agent_name(control_id),  # sanitized agent name for identifier rules
            description=f"Specialist for control {control_id}. Fetches its system instructions from MongoDB by prompt name '{control_id}'",
            instruction=(
                f"{system_instruction}\n\n"
                "Guidance:\n"
                "- The coordinator provides (control_id, app_name).\n"
                "- If the chat has uploaded files, call ingest_uploaded_evidence(files=<the ADK provided uploads>, app_name=<your assigned app_name>) to retrieve from cache.\n"
                "- If no files are available from chat, call load_uploaded_evidence(app_name=<your assigned app_name>) to retrieve from disk.\n"
                "- Select only evidence relevant to your assigned app_name using filenames/content.\n"
                "- Then call analyze_evidence_by_ids(file_ids=[ids of relevant files], system_instructions=this_control_instruction) to process.\n"
                "- Output a concise compliance report for this control and app, with findings and pass/fail per requirement."
            ),
            tools=self.tools + [FunctionTool(self.analyze_evidence_by_ids)],
        )

        # Store the created agent in the pool
        self.control_agents[control_id] = control_agent
        # Track mapping from original control_id to sanitized agent name
        self.control_agent_name_map[control_id] = control_agent.name

        # Attach to coordinator's sub_agents so AutoFlow can route transfer_to_agent
        if not any(a.name == control_agent.name for a in getattr(self.agent, 'sub_agents', [])):
            self.agent.sub_agents.append(control_agent)

        return control_agent

    # Tool: ensure_control_agents
    def ensure_control_agents(self, control_ids: List[str]) -> Dict[str, Dict[str, str]]:
        """Prepare sub-agents for the given control IDs and attach them for transfer routing.
        Returns mapping: control_id -> { status, agent_name, message? }.
        """
        results: Dict[str, Dict[str, str]] = {}
        for cid in control_ids:
            try:
                self._create_control_agent(cid)
                results[cid] = {
                    'status': 'ready',
                    'agent_name': self.control_agent_name_map.get(cid) or self._sanitize_agent_name(cid)
                }
            except Exception as e:
                results[cid] = {
                    'status': 'error',
                    'agent_name': self.control_agent_name_map.get(cid) or self._sanitize_agent_name(cid),
                    'message': str(e)
                }
        return results

    # Tool: load_uploaded_evidence
    def load_uploaded_evidence(self, app_name: Optional[str] = None, dir_path: Optional[str] = None, limit: int = 50) -> List[Dict[str, Any]]:
        """Load uploaded evidence files from a directory and cache contents to avoid token bloat.
        dir_path defaults to env EVIDENCE_UPLOADS_DIR or './uploads'.
        - If app_name is provided, only include files whose filename contains the app_name (case-insensitive).
        Returns: List of manifests { id, filename, size } (no content returned).
        """
        base_dir = Path(dir_path or os.getenv('EVIDENCE_UPLOADS_DIR', 'uploads'))
        if not base_dir.exists() or not base_dir.is_dir():
            logging.warning(f"[load_uploaded_evidence] Uploads directory not found: {base_dir}")
            return []

        allowed = set((os.getenv('ALLOWED_FILE_EXTENSIONS', '') or '').replace(' ', '').split(','))
        results: List[Dict[str, Any]] = []
        all_results: List[Dict[str, Any]] = []
        app_filter = (app_name or '').lower().strip()

        for fp in sorted(base_dir.glob('**/*')):
            if len(results) >= limit:
                break
            if not fp.is_file():
                continue
            ext = fp.suffix.lower().lstrip('.')
            if ext and ext not in allowed:
                continue

            try:
                content_bytes = fp.read_bytes()
                content = base64.b64encode(content_bytes).decode('ascii')
                evid = str(uuid.uuid4())
                self._evidence_cache[evid] = {'filename': fp.name, 'content': content}
                item = {'id': evid, 'filename': fp.name, 'size': len(content)}
                all_results.append(item)
                if not app_filter or app_filter in fp.name.lower():
                    results.append(item)
            except Exception as e:
                logging.warning(f"[load_uploaded_evidence] Failed to process file {fp}: {e}")

        if app_filter and not results:
            results = all_results[:limit]
        logging.info(f"[load_uploaded_evidence] Normalized {len(results)} of {len(all_results)} files")
        return results

    # Tool: ingest_uploaded_evidence
    def ingest_uploaded_evidence(self, files: Optional[List[Any]] = None, app_name: Optional[str] = None, limit: int = 50) -> List[Dict[str, Any]]:
        """Normalize evidence from native ADK Web chat uploads and cache contents.
        Accepts heterogeneous items (filename/name, content/text, data/base64/bytes, path/filepath, meta).
        Returns manifests { id, filename, size } only. Use analyze_evidence_by_ids to process.
        """
        if not files:
            logging.info("[ingest_uploaded_evidence] No files provided; returning empty list.")
            return []

        allowed = set((os.getenv('ALLOWED_FILE_EXTENSIONS', '') or '').replace(' ', '').split(','))
        app_filter = (app_name or '').lower().strip()
        results: List[Dict[str, Any]] = []
        all_results: List[Dict[str, Any]] = []

        def _normalize_one(item: Any) -> Optional[Dict[str, str]]:
            try:
                # Extract filename
                fname = None
                for key in ('filename', 'name', 'file_name'):
                    if isinstance(item, dict) and key in item:
                        fname = str(item[key])
                        break
                # Some ADK wrappers may store metadata under 'meta'
                meta = None
                if not fname and isinstance(item, dict) and isinstance(item.get('meta'), dict):
                    meta = item['meta']
                    for key in ('filename', 'file_name', 'name'):
                        if key in meta:
                            fname = str(meta[key])
                            break
                if not fname:
                    fname = 'uploaded_evidence'

                # Extract content
                content: Optional[str] = None

                # direct string content
                if isinstance(item, dict):
                    for key in ('content', 'text'):
                        if key in item and isinstance(item[key], str):
                            content = item[key]
                            break
                # base64 or bytes
                if content is None and isinstance(item, dict):
                    for key in ('data', 'base64', 'bytes'):
                        if key in item and item[key]:
                            data = item[key]
                            if isinstance(data, (bytes, bytearray)):
                                content = base64.b64encode(data).decode('ascii')
                            else:
                                # assume already base64-encoded string
                                content = str(data)
                            break
                # path on disk
                if content is None and isinstance(item, dict):
                    for key in ('path', 'filepath'):
                        if key in item and item[key]:
                            p = Path(item[key])
                            if p.exists() and p.is_file():
                                ext = p.suffix.lower().lstrip('.')
                                if ext and ext not in allowed:
                                    return None
                                content = base64.b64encode(p.read_bytes()).decode('ascii')
                            break

                if content is None:
                    return None

                evid_id = str(uuid.uuid4())
                self._evidence_cache[evid_id] = {'filename': fname, 'content': content}
                return {'id': evid_id, 'filename': fname, 'size': len(content)}
            except Exception as e:
                logging.warning(f"[ingest_uploaded_evidence] Failed to normalize item: {e}")
                return None

        for it in files[:limit]:
            norm = _normalize_one(it)
            if not norm:
                continue
            all_results.append(norm)
            if not app_filter or app_filter in norm['filename'].lower():
                results.append(norm)

        if app_filter and not results:
            results = all_results[:limit]
        logging.info(f"[ingest_uploaded_evidence] Normalized {len(results)} of {len(all_results)} uploaded items")
        return results

    # Tool: analyze_evidence_by_ids
    def analyze_evidence_by_ids(self, file_ids: List[str], system_instructions: Optional[str] = None) -> Dict[str, Any]:
        """Resolve evidence ids from cache and run analysis. Returns analysis dict.
        Use this instead of passing raw contents through the LLM to avoid token overflows.
        """
        files: List[Dict[str, Any]] = []
        for fid in file_ids:
            item = self._evidence_cache.get(fid)
            if item:
                files.append({'filename': item['filename'], 'content': item['content']})
            else:
                logging.warning(f"[analyze_evidence_by_ids] Missing file id: {fid}")
        return self.mongo_tools.analyze_evidence_files(files, system_instructions)

    # Tool: convert_tabular_to_text_by_ids
    def convert_tabular_to_text_by_ids(
        self,
        file_ids: List[str],
        sheet_name: Optional[str] = None,
        max_rows: int = 1000,
        max_chars: int = 200000,
    ) -> Dict[str, Any]:
        """Convert CSV/Excel evidence into plain text for LLM processing.

        Returns a dict with keys:
        - texts: list of { id, filename, text, rows, cols, note }
        - errors: list of { id, filename, error }
        """
        outputs: List[Dict[str, Any]] = []
        errors: List[Dict[str, str]] = []

        def _decode_with_fallback(data: bytes) -> str:
            for enc in ("utf-8", "utf-16", "utf-16le", "utf-16be", "latin-1"):
                try:
                    return data.decode(enc)
                except Exception:
                    continue
            return data.decode("utf-8", errors="replace")

        for fid in file_ids:
            item = self._evidence_cache.get(fid)
            if not item:
                errors.append({"id": fid, "filename": "", "error": "file id not found in cache"})
                continue
            name = item.get("filename", "")
            content_b64 = item.get("content", "")
            try:
                raw = base64.b64decode(content_b64)
            except Exception as e:
                errors.append({"id": fid, "filename": name, "error": f"invalid base64: {e}"})
                continue

            ext = Path(name).suffix.lower()
            note = ""
            text = ""
            rows = 0
            cols = 0

            try:
                if ext == ".csv":
                    decoded = _decode_with_fallback(raw)
                    reader = csv.reader(decoded.splitlines())
                    lines: List[str] = []
                    for r, row in enumerate(reader):
                        if r >= max_rows:
                            note = f"truncated to {max_rows} rows"
                            break
                        cols = max(cols, len(row))
                        lines.append(" | ".join(str(c) for c in row))
                    text = "\n".join(lines)
                elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"):
                    if _PANDAS_AVAILABLE:
                        try:
                            df = _pd.read_excel(BytesIO(raw), sheet_name=sheet_name or 0, dtype=str, engine=None)
                            if hasattr(df, "to_string"):
                                df = df.head(max_rows)
                                rows, cols = df.shape
                                text = df.to_csv(index=False)
                                if rows >= max_rows:
                                    note = f"truncated to {max_rows} rows"
                        except Exception as e:
                            note = f"pandas excel read failed: {e}"
                    elif _OPENPYXL_AVAILABLE and ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
                        wb = _openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
                        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
                        lines: List[str] = []
                        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if r_idx >= max_rows:
                                note = f"truncated to {max_rows} rows"
                                break
                            row_vals = ["" if v is None else str(v) for v in row]
                            cols = max(cols, len(row_vals))
                            lines.append(" | ".join(row_vals))
                        text = "\n".join(lines)
                    else:
                        raise RuntimeError("Excel parsing unavailable; install pandas or openpyxl")
                else:
                    raise RuntimeError(f"unsupported extension: {ext}")

                if len(text) > max_chars:
                    text = text[:max_chars]
                    note = (note + "; " if note else "") + f"truncated to {max_chars} characters"

                outputs.append({
                    "id": fid,
                    "filename": name,
                    "text": text,
                    "rows": rows,
                    "cols": cols,
                    "note": note,
                })
            except Exception as e:
                errors.append({"id": fid, "filename": name, "error": str(e)})

        return {"texts": outputs, "errors": errors}

    # Tool: extract_outlook_msg_by_ids
    def extract_outlook_msg_by_ids(
        self,
        file_ids: List[str],
        max_chars: int = 200000,
        include_headers: bool = True,
    ) -> Dict[str, Any]:
        """Extracts text and headers from Outlook .msg files in cache.

        Returns a dict with keys:
        - messages: list of { id, filename, subject, sender, to, date, body, note }
        - errors: list of { id, filename, error }
        """
        messages: List[Dict[str, Any]] = []
        errors: List[Dict[str, str]] = []

        for fid in file_ids:
            item = self._evidence_cache.get(fid)
            if not item:
                errors.append({"id": fid, "filename": "", "error": "file id not found in cache"})
                continue
            name = item.get("filename", "")
            ext = Path(name).suffix.lower()
            if ext != ".msg":
                errors.append({"id": fid, "filename": name, "error": "not an Outlook .msg file"})
                continue
            content_b64 = item.get("content", "")
            try:
                raw = base64.b64decode(content_b64)
            except Exception as e:
                errors.append({"id": fid, "filename": name, "error": f"invalid base64: {e}"})
                continue

            try:
                if not _EXTRACT_MSG_AVAILABLE:
                    raise RuntimeError("extract-msg not installed; cannot parse .msg")

                with _extract_msg.openMsg(BytesIO(raw)) as msg:  # type: ignore[attr-defined]
                    subject = (msg.subject or "").strip()
                    sender = (getattr(msg, "sender", "") or getattr(msg, "senderEmail", "") or "").strip()
                    to = (getattr(msg, "to", "") or getattr(msg, "recipients", "") or "").strip()
                    date = str(getattr(msg, "date", "")).strip()
                    body = (msg.body or "").strip()
                    if len(body) > max_chars:
                        body = body[:max_chars]
                        note = f"truncated to {max_chars} characters"
                    else:
                        note = ""
                messages.append({
                    "id": fid,
                    "filename": name,
                    "subject": subject,
                    "sender": sender,
                    "to": to,
                    "date": date,
                    "body": body,
                    "note": note,
                })
            except Exception as e:
                errors.append({"id": fid, "filename": name, "error": str(e)})

        return {"messages": messages, "errors": errors}

    def start(self):
        """Start the NHA Compliance Agent"""
        print("🚀 Starting NHA Compliance Assistant (Non-MCP, Multi-Agent Delegation)")
        print("Ready to route controls to specialized sub-agents and process evidence...")
        return self.agent


def main():
    """Main entry point"""
    try:
        # Initialize the NHA Compliance Agent
        nha_agent = NHAComplianceAgent()
        agent = nha_agent.start()

        print("\n" + "-"*60)
        print("🛡️  NHA COMPLIANCE ASSISTANT (NON-MCP VERSION) READY")
        print("-"*60)
        print("\nTo start validation, provide:")
        print("1. Control ID (required)")
        print("2. Application ID (required)")
        print("3. Jira Ticket ID (optional)")
        print("4. Evidence files (optional)")
        print("\nThe system will then execute the complete validation workflow automatically.")
        print("-"*60)

        return agent

    except Exception as e:
        logger.error(f"Failed to initialize NHA Compliance Agent: {e}")
        raise


# Create root_agent for ADK web server
try:
    root_agent = main()
except Exception as _e:
    # If initialization fails at import time, defer by creating a minimal placeholder agent
    # so the module import doesn't crash environments that introspect 'root_agent'.
    root_agent = LlmAgent(
        model=TachyonAdkClient(model_name=f"openai/{os.getenv('MODEL', 'gemini-2.0-flash')}",),
        name=os.getenv('ROOT_AGENT_NAME', 'NHA_Coordinator'),
        instruction=(
            "Initialization failed; please check server logs."
        ),
        tools=[],
    )
